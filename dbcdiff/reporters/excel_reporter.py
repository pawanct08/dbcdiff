"""
dbcdiff/reporters/excel_reporter.py
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
Export a cantools CAN database to a formatted 4-sheet Excel workbook.

Sheets
------
1. Messages     – one row per message  (Frame ID, Name, DLC, Cycle Time, Sender, Signal Count)
2. Signals      – one row per signal   (full attribute set including unit, min, max, comment)
3. Value Tables – one row per choice entry  (Message, Signal, Raw Value, Description)
4. Nodes        – one row per node     (TX message list + count, RX message list + count)

Usage::

    import cantools
    from dbcdiff.reporters.excel_reporter import write_excel

    db = cantools.database.load_file("vehicle.dbc")
    write_excel(db, "vehicle_matrix.xlsx")

Requires: openpyxl >= 3.1
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Colour palette (Office 2019 "Blue" scheme)
# ---------------------------------------------------------------------------
_H_FILL = PatternFill("solid", fgColor="4472C4")   # header – medium blue
_H_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
_H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

_ODD_FILL = PatternFill("solid", fgColor="DCE6F1")  # alternating row – light blue
_EVN_FILL = PatternFill("solid", fgColor="FFFFFF")  # alternating row – white
_DATA_FONT = Font(size=10, name="Calibri")
_DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _write_header(ws, headers: list[str]) -> None:
    """Write and style the header row (row 1), then freeze it."""
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _H_FILL
        cell.font = _H_FONT
        cell.alignment = _H_ALIGN
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18


def _write_data_row(ws, row_idx: int, values: list, is_odd: bool) -> None:
    """Write a data row with alternating row colour."""
    fill = _ODD_FILL if is_odd else _EVN_FILL
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = fill
        cell.font = _DATA_FONT
        cell.alignment = _DATA_ALIGN


def _autofit_columns(ws, headers: list[str], max_width: int = 60) -> None:
    """Set approximate column widths based on header + data content."""
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        # Start with header width
        col_max = len(str(ws.cell(row=1, column=col_idx).value or "")) + 2
        # Sample every data cell
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    col_max = max(col_max, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = col_max + 2


def _byte_order_label(sig) -> str:
    """Return 'Motorola' or 'Intel' from a cantools signal byte_order."""
    # Works regardless of how cantools formats the ByteOrder enum
    return "Motorola" if "big" in str(sig.byte_order).lower() else "Intel"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_messages(ws, db) -> None:
    """Sheet 1: Messages — one row per CAN message."""
    headers = [
        "Frame ID (Hex)", "Frame ID (Dec)", "Name",
        "DLC (bytes)", "Cycle Time (ms)", "Sender", "Signal Count",
    ]
    _write_header(ws, headers)

    for i, msg in enumerate(sorted(db.messages, key=lambda m: m.frame_id), start=1):
        senders = ", ".join(msg.senders) if msg.senders else ""
        cycle = msg.cycle_time if msg.cycle_time is not None else ""
        row = [
            f"0x{msg.frame_id:03X}",
            msg.frame_id,
            msg.name,
            msg.length,
            cycle,
            senders,
            len(msg.signals),
        ]
        _write_data_row(ws, i + 1, row, is_odd=(i % 2 == 1))

    _autofit_columns(ws, headers)


def _build_signals(ws, db) -> None:
    """Sheet 2: Signals — one row per signal across all messages."""
    headers = [
        "Message", "Frame ID (Hex)", "Signal Name",
        "Start Bit", "Length (bits)", "Byte Order",
        "Scale", "Offset", "Unit", "Min", "Max",
        "Receivers", "Comment",
    ]
    _write_header(ws, headers)

    data_row = 0  # visual counter for alternating colour
    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        for sig in sorted(msg.signals, key=lambda s: s.start):
            scale = sig.scale if sig.scale is not None else 1
            offset = sig.offset if sig.offset is not None else 0
            vmin = sig.minimum if sig.minimum is not None else ""
            vmax = sig.maximum if sig.maximum is not None else ""
            unit = sig.unit or ""
            receivers = ", ".join(sig.receivers) if sig.receivers else ""
            comment = (sig.comment or "").replace("\n", " ").replace("\r", "").strip()

            row = [
                msg.name,
                f"0x{msg.frame_id:03X}",
                sig.name,
                sig.start,
                sig.length,
                _byte_order_label(sig),
                scale,
                offset,
                unit,
                vmin,
                vmax,
                receivers,
                comment,
            ]
            _write_data_row(ws, data_row + 2, row, is_odd=(data_row % 2 == 0))
            data_row += 1

    _autofit_columns(ws, headers)


def _build_value_tables(ws, db) -> None:
    """Sheet 3: Value Tables — one row per signal choice (enumeration) entry.

    Row colours alternate per *message group* rather than per individual row,
    making it easy to see which entries belong to the same message at a glance.
    A Frame ID (Hex) column is included for cross-reference with Sheet 1.
    """
    headers = ["Message", "Frame ID (Hex)", "Signal", "Raw Value", "Description"]
    _write_header(ws, headers)

    data_row = 0
    msg_group = 0   # colour alternates per message that has choices, not per row
    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        has_choices = any(s.choices for s in msg.signals)
        if not has_choices:
            continue
        for sig in sorted(msg.signals, key=lambda s: s.start):
            if not sig.choices:
                continue
            for raw_val, label in sorted(sig.choices.items(), key=lambda kv: kv[0]):
                row = [msg.name, f"0x{msg.frame_id:03X}", sig.name, raw_val, str(label)]
                _write_data_row(ws, data_row + 2, row, is_odd=(msg_group % 2 == 0))
                data_row += 1
        msg_group += 1

    if data_row == 0:
        ws.cell(row=2, column=1, value="(no value tables / enumerations found in this DBC)")

    _autofit_columns(ws, headers)


def _build_nodes(ws, wb, db) -> None:
    """Sheet 4: Nodes — TX/RX message dropdown selectors per network node.

    Each node gets one data row.  The **TX Message** and **RX Message** cells
    contain an in-cell Excel dropdown whose choices are populated from a hidden
    helper sheet (``_NodeLists``).  The user can click the dropdown arrow in any
    node row to browse every message that node transmits or receives without
    leaving Sheet 4.

    Layout
    ------
    | Node Name | TX Message ▼ | TX Count | RX Message ▼ | RX Count |

    The cell is pre-populated with the first (alphabetically) message name;
    selecting a different entry from the dropdown simply shows the chosen name.
    """
    headers = [
        "Node Name",
        "TX Message",   # in-cell dropdown — one TX message at a time
        "TX Count",
        "RX Message",   # in-cell dropdown — one RX message at a time
        "RX Count",
    ]
    _write_header(ws, headers)

    # ---- Build TX / RX lookup maps ----------------------------------------
    tx_map: dict[str, list[str]] = {}
    for msg in db.messages:
        for sender in (msg.senders or []):
            tx_map.setdefault(sender, []).append(msg.name)

    rx_map: dict[str, set[str]] = {}
    for msg in db.messages:
        for sig in msg.signals:
            for recv in (sig.receivers or []):
                rx_map.setdefault(recv, set()).add(msg.name)

    node_names: set[str] = set()
    if db.nodes:
        node_names.update(n.name for n in db.nodes)
    node_names.update(tx_map.keys())
    node_names.update(rx_map.keys())

    if not node_names:
        ws.cell(row=2, column=1, value="(no nodes defined in this DBC)")
        _autofit_columns(ws, headers)
        return

    sorted_nodes = sorted(node_names)

    # ---- Hidden helper sheet: column pairs (TX | RX) per node -------------
    # Node at index i (0-based):  TX column = 2i+1,  RX column = 2i+2
    ws_lists = wb.create_sheet("_NodeLists")
    ws_lists.sheet_state = "hidden"

    # node name → (tx_col_1indexed, tx_count, rx_col_1indexed, rx_count)
    node_meta: dict[str, tuple[int, int, int, int]] = {}
    for idx, name in enumerate(sorted_nodes):
        tx_msgs = sorted(tx_map.get(name, []))
        rx_msgs = sorted(rx_map.get(name, set()))
        tx_col = idx * 2 + 1
        rx_col = idx * 2 + 2
        for r, m in enumerate(tx_msgs, start=1):
            ws_lists.cell(row=r, column=tx_col, value=m)
        for r, m in enumerate(rx_msgs, start=1):
            ws_lists.cell(row=r, column=rx_col, value=m)
        node_meta[name] = (tx_col, len(tx_msgs), rx_col, len(rx_msgs))

    # ---- Write node rows with in-cell dropdowns ---------------------------
    for i, name in enumerate(sorted_nodes, start=1):
        tx_col_n, tx_cnt, rx_col_n, rx_cnt = node_meta[name]
        row_n = i + 1                                   # header occupies row 1
        fill  = _ODD_FILL if (i % 2 == 1) else _EVN_FILL

        tx_msgs = sorted(tx_map.get(name, []))
        rx_msgs = sorted(rx_map.get(name, set()))

        for col, val in [
            (1, name),
            (2, tx_msgs[0] if tx_msgs else ""),   # default = first TX message
            (3, tx_cnt),
            (4, rx_msgs[0] if rx_msgs else ""),   # default = first RX message
            (5, rx_cnt),
        ]:
            c = ws.cell(row=row_n, column=col, value=val)
            c.fill = fill
            c.font = _DATA_FONT
            c.alignment = _DATA_ALIGN

        # TX dropdown — points to this node's TX column in _NodeLists
        if tx_cnt > 0:
            tx_letter = get_column_letter(tx_col_n)
            dv_tx = DataValidation(
                type="list",
                formula1=f"'_NodeLists'!${tx_letter}$1:${tx_letter}${tx_cnt}",
                allow_blank=True,
            )
            ws.add_data_validation(dv_tx)
            dv_tx.add(ws.cell(row=row_n, column=2))

        # RX dropdown — points to this node's RX column in _NodeLists
        if rx_cnt > 0:
            rx_letter = get_column_letter(rx_col_n)
            dv_rx = DataValidation(
                type="list",
                formula1=f"'_NodeLists'!${rx_letter}$1:${rx_letter}${rx_cnt}",
                allow_blank=True,
            )
            ws.add_data_validation(dv_rx)
            dv_rx.add(ws.cell(row=row_n, column=4))

    _autofit_columns(ws, headers)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_excel(db, out_path: str) -> None:
    """Export a cantools database to a formatted 4-sheet Excel workbook.

    Parameters
    ----------
    db:
        A ``cantools.database.Database`` object (already loaded via
        ``cantools.database.load_file()``).
    out_path:
        Destination ``.xlsx`` file path.  The file is created (or
        overwritten) by this function.

    Raises
    ------
    ImportError
        If ``openpyxl`` is not installed.
    """
    wb = Workbook()

    # Remove the default blank sheet that openpyxl always creates
    wb.remove(wb.active)

    ws_msg = wb.create_sheet("Messages")
    _build_messages(ws_msg, db)

    ws_sig = wb.create_sheet("Signals")
    _build_signals(ws_sig, db)

    ws_vt = wb.create_sheet("Value Tables")
    _build_value_tables(ws_vt, db)

    ws_nodes = wb.create_sheet("Nodes")
    _build_nodes(ws_nodes, wb, db)

    wb.save(out_path)
