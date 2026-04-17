"""
dbcdiff.converter
=================
Convert between DBC files and Excel workbooks (.xlsx).

DBC → Excel
-----------
Creates a workbook with three sheets:

* **Messages** – one row per message
  Columns: Frame ID (hex), Name, Length (DLC), Senders, Extended ID, Comment

* **Signals** – one row per signal (linked to its parent message by name)
  Columns: Message Name, Signal Name, Start Bit, Length, Byte Order,
           Signed, Scale, Offset, Min, Max, Unit, Receivers, Comment

* **Nodes** – one row per node defined in the DBC
  Columns: Name, Comment

Excel → DBC
-----------
Reads the same three sheets and reconstructs a DBC file.

Dependencies: cantools, openpyxl
"""

from __future__ import annotations

import os
from typing import TYPE_CHECKING

import cantools
import cantools.database
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:  # pragma: no cover
    from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

_MSG_COLS = [
    "Frame ID (hex)",
    "Name",
    "Length (DLC)",
    "Senders",
    "Extended ID",
    "Comment",
]

_SIG_COLS = [
    "Message Name",
    "Signal Name",
    "Start Bit",
    "Length",
    "Byte Order",
    "Signed",
    "Scale",
    "Offset",
    "Min",
    "Max",
    "Unit",
    "Receivers",
    "Comment",
]

_NODE_COLS = [
    "Name",
    "Comment",
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _write_header(ws: "Worksheet", columns: list[str]) -> None:
    ws.append(columns)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _auto_width(ws: "Worksheet") -> None:
    """Heuristic column auto-width."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _shade_alt_rows(ws: "Worksheet") -> None:
    """Light-blue shading on even data rows for readability."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = _ALT_FILL


# ---------------------------------------------------------------------------
# DBC → Excel
# ---------------------------------------------------------------------------

def dbc_to_excel(dbc_path: str, xlsx_path: str) -> None:
    """Convert *dbc_path* to an Excel workbook saved at *xlsx_path*.

    Parameters
    ----------
    dbc_path:
        Path to the input .dbc file.
    xlsx_path:
        Path where the output .xlsx file will be written.
        Parent directory must already exist.

    Raises
    ------
    FileNotFoundError
        If *dbc_path* does not exist.
    Exception
        Any error raised by cantools or openpyxl is propagated.
    """
    if not os.path.isfile(dbc_path):
        raise FileNotFoundError(f"DBC file not found: {dbc_path}")

    db = cantools.database.load_file(dbc_path)
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ---- Messages sheet ----
    ws_msg: "Worksheet" = wb.create_sheet("Messages")
    _write_header(ws_msg, _MSG_COLS)
    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        senders = ", ".join(msg.senders) if msg.senders else ""
        ws_msg.append([
            f"0x{msg.frame_id:03X}",
            msg.name,
            msg.length,
            senders,
            "Yes" if msg.is_extended_id else "No",
            msg.comment or "",
        ])
    _shade_alt_rows(ws_msg)
    _auto_width(ws_msg)

    # ---- Signals sheet ----
    ws_sig: "Worksheet" = wb.create_sheet("Signals")
    _write_header(ws_sig, _SIG_COLS)
    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        for sig in sorted(msg.signals, key=lambda s: s.name):
            receivers = ", ".join(sig.receivers) if sig.receivers else ""
            byte_order = (
                "big_endian"
                if sig.byte_order == "big_endian"
                else "little_endian"
            )
            ws_sig.append([
                msg.name,
                sig.name,
                sig.start,
                sig.length,
                byte_order,
                "Yes" if sig.is_signed else "No",
                sig.scale,
                sig.offset,
                sig.minimum if sig.minimum is not None else "",
                sig.maximum if sig.maximum is not None else "",
                sig.unit or "",
                receivers,
                sig.comment or "",
            ])
    _shade_alt_rows(ws_sig)
    _auto_width(ws_sig)

    # ---- Nodes sheet ----
    ws_nod: "Worksheet" = wb.create_sheet("Nodes")
    _write_header(ws_nod, _NODE_COLS)
    for node in sorted(db.nodes or [], key=lambda n: n.name):
        ws_nod.append([node.name, node.comment or ""])
    _shade_alt_rows(ws_nod)
    _auto_width(ws_nod)

    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Excel → DBC
# ---------------------------------------------------------------------------

def _parse_bool(value: object, *, true_vals: tuple[str, ...] = ("yes", "true", "1")) -> bool:
    return str(value).strip().lower() in true_vals


def _parse_opt_float(value: object) -> float | None:
    s = str(value).strip()
    if s in ("", "None", "none"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def excel_to_dbc(xlsx_path: str, dbc_path: str) -> None:
    """Convert an Excel workbook at *xlsx_path* back to a DBC file at *dbc_path*.

    The workbook must have at minimum a **Messages** sheet and a **Signals**
    sheet created by :func:`dbc_to_excel` (or manually following the same
    column layout).  The **Nodes** sheet is optional.

    Parameters
    ----------
    xlsx_path:
        Path to the input .xlsx file.
    dbc_path:
        Path where the output .dbc file will be written.

    Raises
    ------
    FileNotFoundError
        If *xlsx_path* does not exist.
    ValueError
        If the workbook is missing required sheets or has malformed data.
    """
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    required = {"Messages", "Signals"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(
            f"Workbook is missing required sheet(s): {', '.join(sorted(missing))}"
        )

    # ---- Read Messages ----
    ws_msg = wb["Messages"]
    msg_rows = list(ws_msg.iter_rows(values_only=True))
    if not msg_rows:
        raise ValueError("Messages sheet is empty")

    # Map header → column index
    hdr = [str(c).strip() if c is not None else "" for c in msg_rows[0]]

    def _col(sheet_hdr: list[str], name: str) -> int:
        try:
            return sheet_hdr.index(name)
        except ValueError:
            raise ValueError(f"Column '{name}' not found in sheet")

    m_frame_id = _col(hdr, "Frame ID (hex)")
    m_name     = _col(hdr, "Name")
    m_length   = _col(hdr, "Length (DLC)")
    m_senders  = _col(hdr, "Senders")
    m_ext      = _col(hdr, "Extended ID")
    m_comment  = _col(hdr, "Comment")

    messages_data: dict[str, dict] = {}
    for row in msg_rows[1:]:
        if not row or row[m_name] is None:
            continue
        fid_raw = str(row[m_frame_id] or "0").strip()
        # Accept both "0x1A2" and plain integers
        try:
            fid = int(fid_raw, 16) if fid_raw.startswith("0x") else int(fid_raw)
        except ValueError:
            fid = 0
        name = str(row[m_name]).strip()
        senders_raw = str(row[m_senders] or "").strip()
        senders = [s.strip() for s in senders_raw.split(",") if s.strip()] if senders_raw else []
        messages_data[name] = {
            "frame_id": fid,
            "name": name,
            "length": int(row[m_length] or 8),
            "senders": senders,
            "is_extended_id": _parse_bool(row[m_ext]),
            "comment": str(row[m_comment] or "").strip() or None,
            "signals": [],
        }

    # ---- Read Signals ----
    ws_sig = wb["Signals"]
    sig_rows = list(ws_sig.iter_rows(values_only=True))
    if sig_rows:
        shdr = [str(c).strip() if c is not None else "" for c in sig_rows[0]]
        s_msg_name  = _col(shdr, "Message Name")
        s_sig_name  = _col(shdr, "Signal Name")
        s_start     = _col(shdr, "Start Bit")
        s_length    = _col(shdr, "Length")
        s_byte_ord  = _col(shdr, "Byte Order")
        s_signed    = _col(shdr, "Signed")
        s_scale     = _col(shdr, "Scale")
        s_offset    = _col(shdr, "Offset")
        s_min       = _col(shdr, "Min")
        s_max       = _col(shdr, "Max")
        s_unit      = _col(shdr, "Unit")
        s_recv      = _col(shdr, "Receivers")
        s_comment   = _col(shdr, "Comment")

        for row in sig_rows[1:]:
            if not row or row[s_msg_name] is None:
                continue
            msg_name = str(row[s_msg_name]).strip()
            if msg_name not in messages_data:
                continue  # orphan signal – skip silently
            recv_raw = str(row[s_recv] or "").strip()
            recv = [r.strip() for r in recv_raw.split(",") if r.strip()] if recv_raw else []
            bo = str(row[s_byte_ord] or "little_endian").strip().lower()
            byte_order = "big_endian" if "big" in bo else "little_endian"
            messages_data[msg_name]["signals"].append({
                "name": str(row[s_sig_name]).strip(),
                "start": int(row[s_start] or 0),
                "length": int(row[s_length] or 1),
                "byte_order": byte_order,
                "is_signed": _parse_bool(row[s_signed]),
                "scale": _parse_opt_float(row[s_scale]) or 1.0,
                "offset": _parse_opt_float(row[s_offset]) or 0.0,
                "minimum": _parse_opt_float(row[s_min]),
                "maximum": _parse_opt_float(row[s_max]),
                "unit": str(row[s_unit] or "").strip() or None,
                "receivers": recv,
                "comment": str(row[s_comment] or "").strip() or None,
            })

    # ---- Read Nodes (optional) ----
    nodes: list = []
    if "Nodes" in wb.sheetnames:
        ws_nod = wb["Nodes"]
        nod_rows = list(ws_nod.iter_rows(values_only=True))
        if len(nod_rows) > 1:
            nhdr = [str(c).strip() if c is not None else "" for c in nod_rows[0]]
            n_name    = nhdr.index("Name") if "Name" in nhdr else 0
            n_comment = nhdr.index("Comment") if "Comment" in nhdr else 1
            for row in nod_rows[1:]:
                if not row or row[n_name] is None:
                    continue
                nodes.append(
                    cantools.database.Node(
                        name=str(row[n_name]).strip(),
                        comment=str(row[n_comment] or "").strip() or None,
                    )
                )

    wb.close()

    # ---- Build cantools DB ----
    messages: list = []
    for md in messages_data.values():
        signals = [
            cantools.database.Signal(
                name=sd["name"],
                start=sd["start"],
                length=sd["length"],
                byte_order=sd["byte_order"],
                is_signed=sd["is_signed"],
                scale=sd["scale"],
                offset=sd["offset"],
                minimum=sd["minimum"],
                maximum=sd["maximum"],
                unit=sd["unit"],
                receivers=sd["receivers"],
                comment=sd["comment"],
            )
            for sd in md["signals"]
        ]
        messages.append(
            cantools.database.Message(
                frame_id=md["frame_id"],
                name=md["name"],
                length=md["length"],
                signals=signals,
                is_extended_id=md["is_extended_id"],
                senders=md["senders"],
                comment=md["comment"],
            )
        )

    db = cantools.database.Database(messages=messages, nodes=nodes)
    db.refresh()

    with open(dbc_path, "w", encoding="utf-8") as fp:
        fp.write(db.as_dbc_string())
